import os
import fitz
from PIL import Image
from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.utils import get_column_letter
import glob

input_folder = r"C:\Users\tamiz\Downloads\oo"
output_dir = "extracted_images"
output_format = "png"
min_width = 100
min_height = 100

if not os.path.exists(output_dir):
    os.makedirs(output_dir)
pdf_files = glob.glob(os.path.join(input_folder, "*.pdf"))

# Create a new workbook
workbook = Workbook()

for file in pdf_files:
    file_name = os.path.splitext(os.path.basename(file))[0]
    pdf_file = fitz.open(file)
    sheet = workbook.create_sheet(title=file_name)
    column_index = 1

    for page_index in range(len(pdf_file)):
        page = pdf_file.load_page(page_index)
        image_list = page.get_images()

        for image_index, image in enumerate(image_list, start=1):
            xref = image[0]
            base_image = pdf_file.extract_image(xref)
            image_data = base_image["image"]
            image_ext = base_image["ext"]
            image_path = os.path.join(output_dir, f"{file_name}_page{page_index + 1}_image{image_index}.{image_ext}")
            with open(image_path, "wb") as img_file:
                img_file.write(image_data)

            img = Image.open(image_path)

            if img.width >= min_width and img.height >= min_height:
                column_letter = get_column_letter(column_index)
                sheet.column_dimensions[column_letter].width = img.width / 6
                img_excel = ExcelImage(image_path)
                img_excel.width *= 1.1
                img_excel.height *= 1.1
                img_excel.anchor = f"{column_letter}1"
                sheet.add_image(img_excel)

                column_index += 1
            else:
                print(f"[-] Skipping image on page {page_index} of {file_name} due to its small size.")

workbook.remove(workbook["Sheet"])

output_excel = os.path.join(input_folder, "extracted_image.xlsx")
workbook.save(output_excel)
print(f"Images extracted and saved to {output_excel} successfully.")
