import os
import fitz
import io
from PIL import Image
from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.utils import get_column_letter

output_dir = "extracted_images"
output_format = "png"
min_width = 100
min_height = 100
if not os.path.exists(output_dir):
    os.makedirs(output_dir)
file = r"C:\Users\tamiz\Downloads\poi_merged.pdf"
pdf_file = fitz.open(file)

# Create a new workbook
workbook = Workbook()
sheet = workbook.active

column_index = 1

for page_index in range(len(pdf_file)):
    page = pdf_file[page_index]
    image_list = page.get_images(full=True)
    if image_list:
        print(f"[+] Found a total of {len(image_list)} images in page {page_index}")
    else:
        print(f"[!] No images found on page {page_index}")
    for image_index, img in enumerate(image_list, start=1):
        xref = img[0]
        base_image = pdf_file.extract_image(xref)
        image_bytes = base_image["image"]
        image_ext = base_image["ext"]
        image = Image.open(io.BytesIO(image_bytes))
        if image.width >= min_width and image.height >= min_height:
            # Save the image to the output directory
            image_path = os.path.join(output_dir, f"image{page_index + 1}_{image_index}.{output_format}")
            image.save(image_path, format=output_format.upper())

            # Insert the image into the Excel sheet in the next available column
            column_letter = get_column_letter(column_index)
            sheet.column_dimensions[column_letter].width = image.width / 6  # Adjust column width based on image size
            img = ExcelImage(image_path)
            img.width *= 1.1  # Adjust the image width for better fit
            img.height *= 1.1  # Adjust the image height for better fit
            img.anchor = f"{column_letter}1"
            sheet.add_image(img)

            column_index += 1  # Move to the next column for the next image
        else:
            print(f"[-] Skipping image {image_index} on page {page_index} due to its small size.")

# Save the workbook with the inserted images
output_excel = "C:/Users/tamiz/Downloads/kl.xlsx"
workbook.save(output_excel)
print(f"Images extracted and saved to {output_excel} successfully.")
